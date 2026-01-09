#!/usr/bin/env python3
"""
Generate a formatted, sortable, filterable Excel workbook from the master schedule.
Matches the formatting style of the Bombay711 pick tracker.

Columns: Date | League | Matchup (Away @ Home) | 1H Score | 2H Score | Final Score
"""

import json
import re  # Added for sanitization
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Paths
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR.parent / "output"
BOX_SCORES_DIR = OUTPUT_DIR / "box_scores"
XLSX_PATH = OUTPUT_DIR / "master_schedule_all_leagues_fixed.xlsx"  # Changed to avoid lock

LEAGUES = ["NBA", "NCAAM", "NFL", "NCAAF"]

# Formatting constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4")
)

# League colors for conditional formatting
LEAGUE_COLORS = {
    "NBA": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),    # Light orange
    "NCAAM": PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid"),  # Light green
    "NFL": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),    # Light blue
    "NCAAF": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),  # Light red
}

# Output columns (simplified)
OUTPUT_HEADERS = ["Date", "League", "Matchup", "1H Away", "1H Home", "2H Away", "2H Home", "Final Away", "Final Home"]


def find_latest_snapshot(league_dir: Path) -> Path:
    """Find the most recent historical snapshot file in a league directory."""
    snapshots = list(league_dir.glob("historical_*.json"))
    if not snapshots:
        return None
    return sorted(snapshots, reverse=True)[0]


def load_all_games() -> list:
    """Load all games from all league snapshots."""
    all_games = []
    
    for league in LEAGUES:
        league_dir = BOX_SCORES_DIR / league
        if not league_dir.exists():
            print(f"  ⚠️  {league} directory not found")
            continue
        
        snapshot = find_latest_snapshot(league_dir)
        if not snapshot:
            print(f"  ⚠️  No snapshot found for {league}")
            continue
        
        with open(snapshot, "r", encoding="utf-8") as f:
            games = json.load(f)
        
        for game in games:
            game["league"] = league
        
        print(f"  ✅ {league}: {len(games)} games from {snapshot.name}")
        all_games.extend(games)
    
    return all_games


def sanitize_string(val):
    """Remove control characters that are invalid in XML."""
    if not isinstance(val, str):
        return val
    # Remove control characters that are invalid in XML (ASCII 0-8, 11-12, 14-31)
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', val)

def transform_game(game: dict) -> dict:
    """Transform raw game data into simplified row format."""
    # Get half scores
    half_scores = game.get("half_scores", {})
    h1 = half_scores.get("H1", {})
    h2 = half_scores.get("H2", {})
    
    # For NFL/NCAAF, check quarter_scores and combine Q1+Q2 for 1H, Q3+Q4 for 2H
    quarter_scores = game.get("quarter_scores", {})
    if quarter_scores and not half_scores:
        q1 = quarter_scores.get("Q1", {})
        q2 = quarter_scores.get("Q2", {})
        q3 = quarter_scores.get("Q3", {})
        q4 = quarter_scores.get("Q4", {})
        h1 = {
            "home": (q1.get("home", 0) or 0) + (q2.get("home", 0) or 0),
            "away": (q1.get("away", 0) or 0) + (q2.get("away", 0) or 0)
        }
        h2 = {
            "home": (q3.get("home", 0) or 0) + (q4.get("home", 0) or 0),
            "away": (q3.get("away", 0) or 0) + (q4.get("away", 0) or 0)
        }
    
    # Build matchup string
    away_team = game.get("away_team_full") or game.get("away_team", "")
    home_team = game.get("home_team_full") or game.get("home_team", "")
    matchup = f"{away_team} @ {home_team}"
    
    return {
        "Date": game.get("date", ""),
        "League": game.get("league", ""),
        "Matchup": sanitize_string(matchup),
        "1H Away": h1.get("away", "") if h1 else "",
        "1H Home": h1.get("home", "") if h1 else "",
        "2H Away": h2.get("away", "") if h2 else "",
        "2H Home": h2.get("home", "") if h2 else "",
        "Final Away": game.get("away_score", ""),
        "Final Home": game.get("home_score", ""),
        "_league_raw": game.get("league", "")  # For coloring
    }


def auto_fit_column_width(ws, col_idx: int, header: str, data_values: list, min_width: int = 8, max_width: int = 50):
    """Calculate optimal column width based on content."""
    # Start with header length
    max_len = len(str(header))
    
    # Check data values
    for val in data_values[:100]:  # Sample first 100 rows
        if val is not None:
            max_len = max(max_len, len(str(val)))
    
    # Add padding and constrain
    width = min(max(max_len + 2, min_width), max_width)
    
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width


def create_formatted_excel():
    """Create formatted Excel workbook."""
    print("\n📊 Generating Formatted Excel Workbook\n")
    print("=" * 60)
    
    # Load data
    raw_games = load_all_games()
    print("=" * 60)
    print(f"\n📊 Total games: {len(raw_games)}")
    
    # Transform to simplified format
    games = [transform_game(g) for g in raw_games]
    
    # Sort by date, then league
    games.sort(key=lambda g: (g.get("Date", ""), g.get("League", "")))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Schedule"
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Write headers
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Collect column data for auto-fit
    column_data = {h: [] for h in OUTPUT_HEADERS}
    
    # Write data rows
    for row_idx, row_data in enumerate(games, 2):
        league = row_data.get("_league_raw", "")
        
        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            value = row_data.get(header, "")
            column_data[header].append(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            
            # Alignment - matchup left, everything else center
            if header == "Matchup":
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN
            
            # League-based row coloring
            if league in LEAGUE_COLORS:
                cell.fill = LEAGUE_COLORS[league]
    
    # Auto-fit column widths
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        if header == "Matchup":
            # Matchup column needs more space
            auto_fit_column_width(ws, col_idx, header, column_data[header], min_width=35, max_width=55)
        elif header == "Date":
            auto_fit_column_width(ws, col_idx, header, column_data[header], min_width=12, max_width=15)
        elif header == "League":
            auto_fit_column_width(ws, col_idx, header, column_data[header], min_width=8, max_width=10)
        else:
            # Score columns
            auto_fit_column_width(ws, col_idx, header, column_data[header], min_width=10, max_width=12)
    
    # Create table for better filtering experience
    # Note: Do not set ws.auto_filter.ref when using a Table, as Tables manage their own filters
    # conflicts can cause Excel repair errors.
    if len(games) > 0:
        last_col = get_column_letter(len(OUTPUT_HEADERS))
        table_ref = f"A1:{last_col}{len(games) + 1}"
        
        # Ensure table name is safe (no spaces, starts with letter)
        table = Table(displayName="MasterScheduleTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.freeze_panes = "A2"
    
    # Summary headers
    summary_headers = ["League", "Games", "Date Range"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Calculate summary stats
    for row_idx, league in enumerate(LEAGUES, 2):
        league_games = [g for g in games if g.get("League") == league]
        dates = [g.get("Date", "") for g in league_games if g.get("Date")]
        date_range = f"{min(dates)} to {max(dates)}" if dates else ""
        
        ws_summary.cell(row=row_idx, column=1, value=league).fill = LEAGUE_COLORS.get(league)
        ws_summary.cell(row=row_idx, column=2, value=len(league_games))
        ws_summary.cell(row=row_idx, column=3, value=date_range)
        
        for col_idx in range(1, 4):
            ws_summary.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_summary.cell(row=row_idx, column=col_idx).alignment = CENTER_ALIGN
    
    # Set summary column widths
    ws_summary.column_dimensions["A"].width = 10
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 28
    
    # Total row
    total_row = len(LEAGUES) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=len(games)).font = Font(bold=True)
    
    # Add metadata sheet
    ws_meta = wb.create_sheet(title="Metadata")
    meta_info = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source", "fetch_completed_boxes.py"),
        ("Total Games", len(games)),
        ("Date Range", f"{min(g.get('Date','') for g in games)} to {max(g.get('Date','') for g in games)}"),
        ("Storage", "metricstrackersgbsv/master-schedules/"),
        ("Columns", "Date, League, Matchup (Away @ Home), 1H/2H/Final Scores"),
    ]
    for row_idx, (key, value) in enumerate(meta_info, 1):
        ws_meta.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_meta.cell(row=row_idx, column=2, value=value)
    ws_meta.column_dimensions["A"].width = 15
    ws_meta.column_dimensions["B"].width = 50
    
    # Save
    wb.save(XLSX_PATH)
    
    print(f"\n💾 Saved: {XLSX_PATH}")
    print(f"   Size: {XLSX_PATH.stat().st_size / 1024:.1f} KB")
    print("\n✨ Features:")
    print("   • Frozen header row")
    print("   • Auto-filter on all columns (sortable/filterable)")
    print("   • Auto-fit column widths with text wrap")
    print("   • League-based row coloring")
    print("   • Simplified columns: Date | League | Matchup | 1H | 2H | Final")
    print("   • Summary sheet with stats")
    
    return XLSX_PATH


if __name__ == "__main__":
    create_formatted_excel()
