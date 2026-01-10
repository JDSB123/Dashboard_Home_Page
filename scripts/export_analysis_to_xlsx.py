import argparse
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles import numbers
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ANALYSIS_DIR = os.path.join(ROOT, 'output', 'analysis')

# Navy blue color for headers
NAVY_BLUE = '1F4E78'
WHITE = 'FFFFFF'
RED = 'FF0000'


def autosize_columns(ws):
    """Auto-size columns based on actual cell content width."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Handle multi-line content
                    lines = str(cell.value).split('\n')
                    cell_len = max(len(line) for line in lines)
                    max_length = max(max_length, cell_len)
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width


def format_workbook(path):
    """Apply professional Excel formatting: headers, body, currency, colors."""
    wb = load_workbook(path)
    ws = wb.active
    
    # Build column index map
    col_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    
    # Stack Date and Time (CST) into single cell with line break
    if 'Date' in col_map and 'Time (CST)' in col_map:
        date_col = col_map['Date']
        time_col = col_map['Time (CST)']
        for row_num in range(2, ws.max_row + 1):
            date_val = ws.cell(row_num, date_col).value
            time_val = ws.cell(row_num, time_col).value
            if date_val and time_val:
                ws.cell(row_num, date_col).value = f"{date_val}\n{time_val}"
                ws.cell(row_num, date_col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Delete Time column and update header
        ws.delete_cols(time_col)
        ws.cell(1, date_col).value = 'Date & Time CST'
        # Rebuild col_map after deletion
        col_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    
    # Apply header formatting: navy blue background, bold white Calibri 9, centered with borders
    header_fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type='solid')
    header_font = Font(name='Calibri', size=9, bold=True, color=WHITE)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='medium', color=WHITE),
        right=Side(style='medium', color=WHITE),
        top=Side(style='medium', color=WHITE),
        bottom=Side(style='medium', color=WHITE)
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Apply body formatting: Calibri 9, centered with borders
    body_font = Font(name='Calibri', size=9)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    # Format currency columns: $#,##0.00
    currency_cols = ['To Risk', 'To Win', 'PnL']
    for col_name in currency_cols:
        if col_name in col_map:
            col_idx = col_map[col_name]
            col_letter = get_column_letter(col_idx)
            for row_num in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row_num}']
                try:
                    val = cell.value
                    if val is not None:
                        cell.value = float(val)
                        cell.number_format = '$#,##0.00'
                        # Red text for negative PnL
                        if col_name == 'PnL' and float(val) < 0:
                            cell.font = Font(name='Calibri', size=9, color=RED)
                except:
                    pass
    
    # Format numeric columns without currency
    numeric_cols = ['Odds']
    for col_name in numeric_cols:
        if col_name in col_map:
            col_idx = col_map[col_name]
            col_letter = get_column_letter(col_idx)
            for row_num in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row_num}']
                try:
                    val = cell.value
                    if val not in (None, '', 'unknown'):
                        cell.value = float(val)
                        cell.number_format = '0'
                except:
                    pass
    
    # Apply Excel Table formatting
    last_col_letter = get_column_letter(len(col_map))
    table_ref = f'A1:{last_col_letter}{ws.max_row}'
    tab = Table(displayName='AnalysisTable', ref=table_ref)
    style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Auto-size columns after all formatting
    autosize_columns(ws)
    
    wb.save(path)


def run(date_str):
    csv_path = os.path.join(ANALYSIS_DIR, f'telegram_analysis_{date_str}.csv')
    if not os.path.exists(csv_path):
        raise FileNotFoundError(csv_path)

    xlsx_path = os.path.join(ANALYSIS_DIR, f'telegram_analysis_{date_str}.xlsx')

    df = pd.read_csv(csv_path)

    # Write to Excel
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='analysis')

    # Apply number formatting
    format_workbook(xlsx_path)
    print(xlsx_path)


def cli():
    p = argparse.ArgumentParser()
    p.add_argument('--date', '-d', required=True)
    args = p.parse_args()
    out = run(args.date)
    print('Saved:', out)


if __name__ == '__main__':
    cli()
